import os
from openpyxl import load_workbook
from scraper import EprocScraper

def run_tests(planilha_path=None):
    if planilha_path is None:
        planilha_path = os.path.join(os.path.dirname(__file__), "..", "platform_manager", "base_clientes.xlsx")
    
    if not os.path.exists(planilha_path):
        print(f"❌ Erro: Planilha não encontrada em {planilha_path}")
        return
        
    # Abriremos a planilha sem 'read_only' para podermos injetar os dados de volta
    wb = load_workbook(filename=planilha_path)
    sheet = wb.active
    
    # Preparando cabeçalhos para os resultados
    sheet.cell(row=1, column=3, value="Status Consulta")
    sheet.cell(row=1, column=4, value="Classe da Ação")
    sheet.cell(row=1, column=5, value="Data da Movimentação")
    sheet.cell(row=1, column=6, value="Descrição da Movimentação")
    sheet.cell(row=1, column=7, value="Triagem")
    
    
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) <= 1:
        print("❌ A planilha está vazia ou contém apenas o cabeçalho.")
        return
        
    dados = rows[1:]
    
    print(f"📋 Inciando processamento em lote com {len(dados)} clientes da planilha.")
    
    resultados_consolidados = []
    
    # Rodando em modo normal (visível) pois o Cloudflare exige janela no monitor
    with EprocScraper(headless=False) as scraper:
        for idx, linha in enumerate(dados):
            cpf = linha[0]
            processo_esperado = linha[1] if len(linha) > 1 else None
            
            if not cpf:
                continue
                
            print(f"\n==================================================")
            print(f"[{idx+1}/{len(dados)}] 🔍 Consultando CPF: {cpf}")
            if processo_esperado:
                print(f"Modo: 🎯 MONITORAMENTO (Processo Esperado: {processo_esperado})")
            else:
                print(f"Modo: 📡 DESCOBERTA (Buscando processos de interesse)")
            print(f"==================================================")
            
            # Valor padrão caso falhe
            resultado = {
                "Iteração": idx + 1,
                "CPF": cpf,
                "Processo": processo_esperado,
                "Status": "ERRO/NÃO CONSULTADO",
                "Triagem": "-"
            }
            
            try:
                # Agora passamos o processo esperado para que ele possa ser clicado na lista
                scraper.consultar_processo(cpf=cpf, processo_esperado=processo_esperado)
            except Exception as e:
                print(f"❌ Erro ao consultar {cpf}: {e}")
                resultados_consolidados.append(resultado)
                sheet.cell(row=idx + 2, column=3, value="ERRO NA BUSCA")
                continue
            
            # O scraper salva o HTML atual no arquivo 'page_source.html'
            if os.path.exists("page_source.html"):
                with open("page_source.html", "r", encoding="utf-8") as f:
                    html = f.read()
                    
                    
                # BIFURCAÇÃO DA ARQUITETURA
                linha_planilha = idx + 2
                
                if processo_esperado:
                    # MODO MONITORAMENTO (Comportamento antigo)
                    if processo_esperado in html:
                        print(f"✅ SUCESSO: Processo {processo_esperado} ENCONTRADO na página!")
                        
                        import re
                        
                        # Extrair Classe da Ação usando Regex direto no HTML
                        acao_str = ""
                        match_acao = re.search(r'Classe da a.ão:.*?<[^>]+>([^<]+)', html, re.IGNORECASE | re.DOTALL)
                        if not match_acao:
                            # Fallback
                            match_acao = re.search(r'Classe da a.ão:\s*([^<]+)', html, re.IGNORECASE)
                        
                        if match_acao:
                            acao_str = match_acao.group(1).strip()
                                
                        # Extrair a última movimentação (Data e Descrição) usando Regex
                        movimentacao_str = ""
                        data_str = ""
                        
                        # Procura a primeira TD que tenha uma data e captura ela e o conteúdo da TD logo em seguida
                        match_mov = re.search(r'<td[^>]*>\s*(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})\s*</td>\s*<td[^>]*>(.*?)</td>', html, re.IGNORECASE | re.DOTALL)
                        
                        if match_mov:
                            data_str = match_mov.group(1).strip()
                            # A descrição pode conter tags HTML dentro (como <a>), então removemos as tags HTML internas para ficar limpo
                            desc_bruta = match_mov.group(2).strip()
                            movimentacao_str = re.sub(r'<[^>]+>', '', desc_bruta).strip()
                        
                        print(f"📌 Classe da Ação Extraída: {acao_str if acao_str else 'NÃO ENCONTRADA'}")
                        print(f"📌 Última Movimentação (Data): {data_str if data_str else 'NÃO ENCONTRADA'}")
                        print(f"📌 Última Movimentação (Desc): {movimentacao_str if movimentacao_str else 'NÃO ENCONTRADA'}")
                        
                        alerta = "-"
                        if acao_str and movimentacao_str:
                            alerta = scraper.aplicar_triagem(acao_str, movimentacao_str)
                            print(f"🤖 Resultado da Triagem: {alerta}")
                        
                        resultado["Status"] = "ENCONTRADO"
                        resultado["Triagem"] = alerta
                        
                        # Salva os dados na planilha
                        sheet.cell(row=linha_planilha, column=3, value="ENCONTRADO")
                        sheet.cell(row=linha_planilha, column=4, value=acao_str if acao_str else "NÃO ENCONTRADA")
                        sheet.cell(row=linha_planilha, column=5, value=data_str if data_str else "NÃO ENCONTRADA")
                        sheet.cell(row=linha_planilha, column=6, value=movimentacao_str if movimentacao_str else "NÃO ENCONTRADA")
                        sheet.cell(row=linha_planilha, column=7, value=alerta)
                        
                    else:
                        print(f"❌ FALHA: Processo {processo_esperado} NÃO encontrado na página.")
                        resultado["Status"] = "NÃO ENCONTRADO NA LISTA"
                        sheet.cell(row=linha_planilha, column=3, value="NÃO ENCONTRADO NA LISTA")
                        
                else:
                    # MODO DESCOBERTA (Só tem o CPF)
                    html_upper = html.upper()
                    classes_perigosas = [
                        "BUSCA E APREENSÃO", "EXECUÇÃO DE TÍTULO EXTRAJUDICIAL",
                        "AÇÃO MONITÓRIA", "MONITÓRIA", "AÇÃO DE COBRANÇA", "COBRANÇA"
                    ]
                    
                    encontrou_perigo = False
                    for classe in classes_perigosas:
                        if classe in html_upper:
                            encontrou_perigo = True
                            print(f"⚠️ DESCOBERTA: Possível '{classe}' identificada na lista do CPF!")
                            break
                            
                    if encontrou_perigo:
                        resultado["Status"] = "⚠️ DESCOBERTA: PROCESSO IDENTIFICADO"
                        resultado["Triagem"] = "🚨 REQUER VALIDAÇÃO DO CS"
                        
                        sheet.cell(row=linha_planilha, column=3, value="PROCESSO IDENTIFICADO")
                        sheet.cell(row=linha_planilha, column=7, value="🚨 REQUER VALIDAÇÃO DO CS")
                    else:
                        print(f"✅ DESCOBERTA: Nenhum processo perigoso encontrado na lista para este CPF.")
                        resultado["Status"] = "NENHUM PROCESSO ENCONTRADO"
                        resultado["Triagem"] = "✅ LIMPO"
                        
                        sheet.cell(row=linha_planilha, column=3, value="NENHUM PROCESSO ENCONTRADO")
                        sheet.cell(row=linha_planilha, column=7, value="✅ LIMPO")
                        
            else:
                print("❌ FALHA: Arquivo page_source.html não foi gerado.")
                sheet.cell(row=idx + 2, column=3, value="ERRO HTML")
                
            resultados_consolidados.append(resultado)

    # Imprimir o Relatório Consolidado no final
    print("\n\n" + "="*80)
    print("📊 RELATÓRIO CONSOLIDADO DAS CONSULTAS")
    print("="*80)
    print(f"{'#':<4} | {'CPF':<15} | {'PROCESSO':<26} | {'STATUS':<22} | {'TRIAGEM'}")
    print("-"*80)
    for res in resultados_consolidados:
        print(f"{res['Iteração']:<4} | {res['CPF']:<15} | {res['Processo']:<26} | {res['Status']:<22} | {res['Triagem']}")
    print("="*80)
    
    # Salvar a planilha com os dados atualizados
    try:
        wb.save(planilha_path)
        print(f"\n💾 Planilha salva com sucesso com as informações extraídas!")
    except PermissionError:
        print(f"\n❌ ERRO FATAL: O robô não conseguiu salvar a planilha porque ela está aberta no Excel. Por favor, feche a planilha e rode o robô novamente.")

if __name__ == "__main__":
    run_tests()
