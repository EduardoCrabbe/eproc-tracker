import os
import requests
from openpyxl import load_workbook

def validar_contrato_planilha(caminho: str) -> bool:
    """Verifica se o .xlsx possui o contrato mínimo de dados."""
    if not os.path.exists(caminho):
        print(f"❌ [ERRO] Planilha não encontrada: {caminho}")
        return False
    
    try:
        wb = load_workbook(filename=caminho, read_only=True)
        sheet = wb.active
        cabecalho = next(sheet.iter_rows(max_row=1, values_only=True), [])
        colunas = [str(c).strip().upper() for c in cabecalho if c]
        
        obrigatorias = ["CPF", "NÚMERO DO PROCESSO"]
        if all(col in colunas for col in obrigatorias):
            print("✅ [OK] Contrato de dados validado.")
            return True
        else:
            print(f"❌ [ERRO] Colunas faltando. Necessárias: {obrigatorias}")
            return False
    except Exception as e:
        print(f"💥 [ERRO] Falha ao ler planilha: {e}")
        return False

def testar_conexao_eproc():
    """Verifica se o portal Eproc está respondendo."""
    try:
        # URL exemplo (ajustaremos conforme o subdomínio específico do seu Eproc)
        response = requests.get("https://esaj.tjsp.jus.br", timeout=10)
        if response.status_code == 200:
            print("✅ [OK] Eproc SP online.")
            return True
        else:
            print(f"⚠️ [ALERTA] Eproc retornou status: {response.status_code}")
            return False
    except Exception as e:
        print(f"❌ [ERRO] Sem conexão com Eproc: {e}")
        return False

if __name__ == "__main__":
    print("--- 🛡️ Iniciando Pre-flight Check ---")
    planilha = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "platform_manager", "base_clientes.xlsx") # Nome padrão do nosso arquivo
    if validar_contrato_planilha(planilha) and testar_conexao_eproc():
        print("🚀 Ambiente pronto para automação!")
    else:
        print("🛑 Pre-flight falhou. Corrija o ambiente antes de rodar o robô.")
