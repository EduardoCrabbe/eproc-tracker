from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
import os
import sys
import threading

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EPROC_DIR = os.path.join(BASE_DIR, "eproc_consultant")
UPLOADS_DIR = os.path.join(BASE_DIR, "platform_manager", "uploads")
sys.path.append(EPROC_DIR)

if not os.path.exists(UPLOADS_DIR):
    os.makedirs(UPLOADS_DIR)

try:
    from test_scraper import run_tests
except ImportError:
    pass

app = FastAPI(title="Eproc Tracker API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Estado global do robô por usuário
robo_states = {}

class ClienteNovo(BaseModel):
    cpf: str
    processo: str = ""

class CRMAtendimento(BaseModel):
    codigo_dj: str

class CRMQuitar(BaseModel):
    codigo_dj: str
    data_boleto: str
    data_pagamento: str

def get_user_file(username: str):
    if not username:
        raise HTTPException(status_code=400, detail="Username is required")
    return os.path.join(UPLOADS_DIR, f"base_{username}.xlsx")

@app.get("/api/clientes")
def listar_clientes(user: str):
    planilha_path = get_user_file(user)
    if not os.path.exists(planilha_path):
        return {"clientes": []}
    
    wb = load_workbook(planilha_path, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) <= 1:
        return {"clientes": []}
        
    clientes = []
    for idx, row in enumerate(rows[1:]):
        cpf = row[0] if len(row) > 0 else ""
        if not cpf:
            continue
            
        clientes.append({
            "id": idx + 2,
            "cpf": cpf,
            "processo": row[1] if len(row) > 1 else "",
            "status": row[2] if len(row) > 2 else "",
            "classe": row[3] if len(row) > 3 else "",
            "data": row[4] if len(row) > 4 else "",
            "desc": row[5] if len(row) > 5 else "",
            "triagem": row[6] if len(row) > 6 else ""
        })
    return {"clientes": clientes}

@app.post("/api/clientes")
def adicionar_cliente(cliente: ClienteNovo, user: str):
    planilha_path = get_user_file(user)
    if not os.path.exists(planilha_path):
        # Cria a planilha se não existir
        wb = Workbook()
        sheet = wb.active
        sheet.append(["CPF", "Processo", "Status Consulta", "Classe da Ação", "Data da Movimentação", "Descrição da Movimentação", "Triagem"])
    else:
        wb = load_workbook(planilha_path)
        sheet = wb.active
        
    nova_linha = sheet.max_row + 1
    sheet.cell(row=nova_linha, column=1, value=cliente.cpf)
    sheet.cell(row=nova_linha, column=2, value=cliente.processo)
    
    wb.save(planilha_path)
    return {"message": "Cliente inserido com sucesso!"}

@app.post("/api/upload")
async def upload_planilha(user: str = Query(...), file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Envie apenas arquivos Excel (.xlsx)")
        
    planilha_path = get_user_file(user)
    conteudo = await file.read()
    with open(planilha_path, "wb") as f:
        f.write(conteudo)
        
    return {"message": "Planilha atualizada com sucesso!"}

def rodar_robo_em_background(user: str, planilha_path: str):
    global robo_states
    robo_states[user] = {"is_running": True, "message": "Robô em execução..."}
    try:
        run_tests(planilha_path)
        robo_states[user] = {"is_running": False, "message": "Varredura concluída com sucesso!"}
    except Exception as e:
        robo_states[user] = {"is_running": False, "message": f"Erro na execução: {e}"}

@app.post("/api/scan")
def disparar_varredura(user: str):
    global robo_states
    state = robo_states.get(user, {"is_running": False})
    if state.get("is_running"):
        return {"message": "O robô já está rodando!"}
        
    planilha_path = get_user_file(user)
    if not os.path.exists(planilha_path):
        raise HTTPException(status_code=404, detail="Nenhuma planilha encontrada para varrer.")
        
    thread = threading.Thread(target=rodar_robo_em_background, args=(user, planilha_path))
    thread.start()
    return {"message": "Varredura iniciada em background."}

@app.get("/api/status")
def status_robo(user: str):
    global robo_states
    return robo_states.get(user, {"is_running": False, "message": "Parado"})

# ========================================================
# ROTAS DO CRM (ABA AREA CS / MEUS CLIENTES)
# ========================================================
import time

def get_crm_file(username: str):
    return os.path.join(UPLOADS_DIR, f"atendimentos_{username}.xlsx")

def formatar_nome(nome_completo: str):
    if not nome_completo:
        return ""
    partes = str(nome_completo).strip().split()
    if len(partes) >= 2:
        return f"{partes[0]} {partes[1]}"
    return partes[0] if partes else ""

@app.post("/api/areacs/upload")
async def upload_planilha_crm(user: str = Query(...), file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Envie apenas arquivos Excel (.xlsx)")
        
    caminho_temporario = get_crm_file(f"temp_{user}")
    caminho_final = get_crm_file(user)
    
    conteudo = await file.read()
    with open(caminho_temporario, "wb") as f:
        f.write(conteudo)
        
    try:
        wb_temp = load_workbook(caminho_temporario, data_only=True)
        sheet_temp = wb_temp.active
        
        wb_final = Workbook()
        sheet_final = wb_final.active
        
        # Cabeçalhos do CRM
        headers = ["Código DJ", "Cliente", "UF", "Tipo de contrato", "Processo?", "Criticidade", "Contatos", "Status", "Último Contato"]
        sheet_final.append(headers)
        
        for idx, row in enumerate(sheet_temp.iter_rows(values_only=True)):
            if idx == 0:
                continue # Pula cabeçalho
                
            codigo_dj = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
            cliente_cru = row[1] if len(row) > 1 and row[1] is not None else ""
            tipo_contrato = row[2] if len(row) > 2 and row[2] is not None else "Veículo"
            processos = row[3] if len(row) > 3 and row[3] is not None else "Não"
            atendimento_inicial = row[4] if len(row) > 4 and row[4] is not None else 0
            
            if not codigo_dj:
                continue
                
            cliente_formatado = formatar_nome(cliente_cru)
            uf = "SP" # Default, poderia ser extraido de outro campo
            
            # Limpando atendimento inicial
            try:
                atendimentos = int(atendimento_inicial)
            except:
                atendimentos = 0
                
            # Default values for new CRM logic
            criticidade = "Regular"
            status = "Ativo"
            ultimo_contato = ""
            
            sheet_final.append([codigo_dj, cliente_formatado, uf, tipo_contrato, processos, criticidade, atendimentos, status, ultimo_contato])
            
        wb_final.save(caminho_final)
        os.remove(caminho_temporario)
        return {"message": "Planilha CRM importada e formatada com sucesso!"}
    except Exception as e:
        if os.path.exists(caminho_temporario):
            os.remove(caminho_temporario)
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilha: {e}")

@app.get("/api/areacs/clientes")
def listar_clientes_crm(user: str):
    planilha_path = get_crm_file(user)
    if not os.path.exists(planilha_path):
        return {"clientes": []}
    
    try:
        wb = load_workbook(planilha_path, data_only=True)
        sheet = wb.active
        
        clientes = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                continue
                
            clientes.append({
                "id": str(row[0]),
                "nome": str(row[1]),
                "uf": str(row[2]),
                "contrato": str(row[3]),
                "processos": str(row[4]),
                "criticidade": str(row[5]),
                "contatos": int(row[6] if row[6] is not None else 0),
                "status": str(row[7]),
                "ultimoContato": float(row[8]) if row[8] else None
            })
        return {"clientes": clientes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler CRM: {e}")

@app.post("/api/areacs/atendimento/{codigo_dj}")
def registrar_atendimento(codigo_dj: str, user: str):
    planilha_path = get_crm_file(user)
    if not os.path.exists(planilha_path):
        raise HTTPException(status_code=404, detail="Planilha não encontrada.")
        
    try:
        wb = load_workbook(planilha_path)
        sheet = wb.active
        
        atualizado = False
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_idx, column=1).value) == str(codigo_dj):
                contatos = sheet.cell(row=row_idx, column=7).value
                contatos = int(contatos) if contatos is not None else 0
                if contatos < 6:
                    sheet.cell(row=row_idx, column=7, value=contatos + 1)
                    sheet.cell(row=row_idx, column=9, value=time.time() * 1000) # JS usa ms
                atualizado = True
                break
                
        if atualizado:
            wb.save(planilha_path)
            return {"message": "Atendimento registrado."}
        else:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/areacs/quitar/{codigo_dj}")
def quitar_contrato(codigo_dj: str, payload: CRMQuitar, user: str):
    planilha_path = get_crm_file(user)
    if not os.path.exists(planilha_path):
        raise HTTPException(status_code=404, detail="Planilha não encontrada.")
        
    try:
        wb = load_workbook(planilha_path)
        sheet = wb.active
        
        atualizado = False
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_idx, column=1).value) == str(codigo_dj):
                sheet.cell(row=row_idx, column=8, value="Quitado")
                atualizado = True
                break
                
        if atualizado:
            wb.save(planilha_path)
            return {"message": "Contrato quitado."}
        else:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/areacs/cliente/{codigo_dj}")
def atualizar_cliente(codigo_dj: str, user: str, campo: str = Query(...), valor: str = Query(...)):
    planilha_path = get_crm_file(user)
    if not os.path.exists(planilha_path):
        raise HTTPException(status_code=404, detail="Planilha não encontrada.")
        
    col_map = {"criticidade": 6, "processos": 5}
    if campo not in col_map:
        raise HTTPException(status_code=400, detail="Campo inválido.")
        
    try:
        wb = load_workbook(planilha_path)
        sheet = wb.active
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_idx, column=1).value) == str(codigo_dj):
                sheet.cell(row=row_idx, column=col_map[campo], value=valor)
                break
        wb.save(planilha_path)
        return {"message": "Cliente atualizado."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
